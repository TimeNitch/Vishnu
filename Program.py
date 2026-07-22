import json
import time
import tkinter as tk
import threading
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from datetime import datetime
from pathlib import Path

from ahk import AHK
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import io
import urllib.request

from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage


# ============================================================
# การตั้งค่า
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

EXCEL_FILE = BASE_DIR / "Data_Vishnu.xlsx"
TEMP_FILE = BASE_DIR / "temp_vishnu_data.txt"
TEMP_PROFILES_DIR = BASE_DIR / "temp_profiles"

LOCAL_SERVER_HOST = "127.0.0.1"
LOCAL_SERVER_PORT = 8765

SHEET_NAME = "Data"

AHK_EXE_PATH = r"C:\Program Files\AutoHotkey\v2\AutoHotkey.exe"

CONSOLE_OPEN_DELAY = 1
JAVASCRIPT_RUN_DELAY = 1
RESULT_WAIT_TIMEOUT = 900
RESULT_POLL_INTERVAL = 0.5

# จำนวนคนสูงสุดที่ต้องการทดสอบ
MAX_PROFILES = 1000

# จำนวนข้อทั้งหมด ใช้สร้างค่า เช่น 1/56
TOTAL_QUESTIONS = 56

# ลบไฟล์ temp หลังบันทึก Excel สำเร็จหรือไม่
DELETE_TEMP_AFTER_SUCCESS = True


# ============================================================
# JavaScript สำหรับดึงข้อมูล
# ============================================================

JAVASCRIPT = r"""
(async () => {
    const MAX_PROFILES = __MAX_PROFILES__;
    const TOTAL_QUESTIONS = __TOTAL_QUESTIONS__;
    const cleanText = (value) => {
        return String(value ?? "")
            .replace(/\s+/g, " ")
            .trim();
    };

    const sleep = (milliseconds) => {
        return new Promise((resolve) => setTimeout(resolve, milliseconds));
    };

    const postToPython = async (path, payload) => {
        const response = await fetch(
            `http://__SERVER_HOST__:__SERVER_PORT__${path}`,
            {
                method: "POST",
                headers: {
                    "Content-Type": "application/json"
                },
                body: JSON.stringify(payload)
            }
        );

        if (!response.ok) {
            const errorText = await response.text();

            throw new Error(
                `ส่งข้อมูลไป Python ไม่สำเร็จ: ${response.status} ${errorText}`
            );
        }

        return response.json();
    };

    const expectedFields = [
        "ชื่อเล่น",
        "จังหวัด",
        "โรงเรียนเดิม",
        "ช่องทางติดต่อ",
        "อาหารที่ชอบ",
        "สิ่งที่ชอบ",
        "ความสามารถ/งานอดิเรก",
        "สายการเรียน",
        "แท็กบุคลิก"
    ];

    const ignoredFields = new Set([
        "Profile Picture",
        "Profile Picture URL",
        "Page Title",
        "Source URL",
        "Saved At"
    ]);

    const normalizeForCompare = (value) => {
        return cleanText(value).toLocaleLowerCase("th-TH");
    };

    const extractCurrentProfileData = () => {
        let profileImage = document.querySelector(
            'picture img[alt="ข้อใด คือคุณสมบัติของเพื่อนคนนี้"]'
        );

        if (!profileImage) {
            profileImage = Array.from(
                document.querySelectorAll("picture img")
            ).find((image) => {
                const alt = cleanText(image.getAttribute("alt"));

                return (
                    alt.includes("คุณสมบัติของเพื่อน") ||
                    alt.includes("เพื่อนคนนี้")
                );
            });
        }

        if (!profileImage) {
            profileImage = document.querySelector("picture img");
        }

        const informationList = Array.from(
            document.querySelectorAll("ul")
        ).find((list) => {
            const text = cleanText(list.textContent);

            const matchCount = expectedFields.filter(
                (field) => text.includes(field)
            ).length;

            return matchCount >= 2;
        });

        if (!informationList) {
            throw new Error("ไม่พบรายการข้อมูลของคนปัจจุบัน");
        }

        const data = {
            "Profile Picture": profileImage
                ? (
                    profileImage.currentSrc ||
                    profileImage.src ||
                    profileImage.getAttribute("src") ||
                    ""
                )
                : ""
        };

        const items = Array.from(
            informationList.querySelectorAll(":scope > li")
        );

        for (const item of items) {
            const paragraphs = Array.from(
                item.querySelectorAll(":scope > p")
            );

            if (paragraphs.length < 2) {
                continue;
            }

            const fieldName = cleanText(paragraphs[0].textContent);

            const fieldValue = paragraphs
                .slice(1)
                .map((paragraph) => cleanText(paragraph.textContent))
                .filter(Boolean)
                .join(" | ");

            if (!fieldName) {
                continue;
            }

            if (data[fieldName] && fieldValue) {
                data[fieldName] += " | " + fieldValue;
            } else {
                data[fieldName] = fieldValue;
            }
        }

        data["Page Title"] = document.title || "";
        data["Source URL"] = location.href;

        return data;
    };

    const buildExtractedValues = (data) => {
        const values = new Set();

        for (const [fieldName, fieldValue] of Object.entries(data)) {
            if (
                ignoredFields.has(fieldName) ||
                typeof fieldValue !== "string"
            ) {
                continue;
            }

            const separatedValues = fieldValue
                .split("|")
                .map((value) => normalizeForCompare(value))
                .filter(Boolean);

            for (const value of separatedValues) {
                values.add(value);
            }

            const completeValue = normalizeForCompare(fieldValue);

            if (completeValue) {
                values.add(completeValue);
            }
        }

        return values;
    };

    const findButtonByText = (text) => {
        return Array.from(
            document.querySelectorAll("button")
        ).find((button) => {
            return cleanText(button.textContent) === text;
        });
    };

    const getAnswerButtons = () => {
        return Array.from(
            document.querySelectorAll('button[type="button"]')
        ).filter((button) => {
            return Boolean(
                button.querySelector(
                    'img[src*="background-image-card"]'
                )
            );
        });
    };

    const waitForStartPageOrFinish = async () => {
        for (let attempt = 0; attempt < 80; attempt += 1) {
            const startButton = findButtonByText("เริ่มต้นตอบคำถาม");

            if (startButton) {
                return {
                    finished: false,
                    startButton
                };
            }

            const answerButtons = getAnswerButtons();
            const submitButton = findButtonByText("ส่งคำตอบ");

            if (answerButtons.length === 0 && !submitButton) {
                const informationListExists = Array.from(
                    document.querySelectorAll("ul")
                ).some((list) => {
                    const text = cleanText(list.textContent);

                    return expectedFields.filter(
                        (field) => text.includes(field)
                    ).length >= 2;
                });

                if (!informationListExists && attempt >= 8) {
                    return {
                        finished: true,
                        startButton: null
                    };
                }
            }

            await sleep(250);
        }

        throw new Error(
            'รอหน้าข้อมูลถัดไปหรือปุ่ม "เริ่มต้นตอบคำถาม" เกินกำหนด'
        );
    };

    const waitForAnswerButtons = async () => {
        for (let attempt = 0; attempt < 40; attempt += 1) {
            const buttons = getAnswerButtons();

            if (buttons.length >= 8) {
                return buttons;
            }

            await sleep(100);
        }

        throw new Error("รอตัวเลือกคำตอบ 8 ช่องเกินกำหนด");
    };

    const waitForSubmitButton = async () => {
        for (let attempt = 0; attempt < 20; attempt += 1) {
            const submitButton = findButtonByText("ส่งคำตอบ");

            if (
                submitButton &&
                !submitButton.disabled &&
                submitButton.getAttribute("aria-busy") !== "true"
            ) {
                return submitButton;
            }

            await sleep(100);
        }

        throw new Error('ปุ่ม "ส่งคำตอบ" ยังไม่พร้อมใช้งาน');
    };

    let savedProfileCount = 0;
    let internalQuestionNumber = 1;

    while (true) {
        const pageState = await waitForStartPageOrFinish();

        if (pageState.finished) {
            console.log("VISHNU_ALL_QUESTIONS_COMPLETED");
            break;
        }

        const currentData = extractCurrentProfileData();
        const extractedValues = buildExtractedValues(currentData);

        console.log(
            `VISHNU_PROFILE_DATA_QUESTION_${internalQuestionNumber}`,
            currentData
        );

        const profilePayload = {
            "Question": `${internalQuestionNumber}/${TOTAL_QUESTIONS}`,
            ...currentData,
            "Collected At": new Date().toISOString()
        };

        const saveResponse = await postToPython(
            "/profile",
            profilePayload
        );

        savedProfileCount = saveResponse.saved_count;

        console.log(
            `VISHNU_PROFILE_SAVED_${savedProfileCount}`,
            profilePayload
        );

        const startButton = pageState.startButton;

        if (
            startButton.disabled ||
            startButton.getAttribute("aria-busy") === "true"
        ) {
            throw new Error(
                `ข้อ ${internalQuestionNumber}: ` +
                'ปุ่ม "เริ่มต้นตอบคำถาม" ยังไม่พร้อมใช้งาน'
            );
        }

        startButton.click();

        console.log(
            `VISHNU_START_BUTTON_CLICKED_QUESTION_${internalQuestionNumber}`
        );

        await sleep(500);

        const answerButtons = await waitForAnswerButtons();

        let matchedCount = 0;
        const matchedAnswers = [];

        for (const button of answerButtons) {
            if (matchedCount >= 4) {
                break;
            }

            const image = button.querySelector("img[alt]");

            const optionText = cleanText(
                image?.getAttribute("alt") ||
                button.textContent
            );

            const normalizedOption = normalizeForCompare(optionText);

            if (
                !normalizedOption ||
                !extractedValues.has(normalizedOption) ||
                button.disabled ||
                button.getAttribute("aria-busy") === "true"
            ) {
                continue;
            }

            button.click();
            matchedCount += 1;
            matchedAnswers.push(optionText);

            await sleep(25);
        }

        console.log(
            `VISHNU_MATCHED_ANSWERS_QUESTION_${internalQuestionNumber}`,
            matchedAnswers
        );

        if (matchedCount !== 4) {
            throw new Error(
                `ข้อ ${internalQuestionNumber}: ` +
                `จับคู่คำตอบได้ ${matchedCount} ช่อง แต่ต้องครบ 4 ช่อง`
            );
        }

        const submitButton = await waitForSubmitButton();

        submitButton.click();

        console.log(
            `VISHNU_SUBMIT_BUTTON_CLICKED_QUESTION_${internalQuestionNumber}`
        );

        internalQuestionNumber += 1;

        if (savedProfileCount >= MAX_PROFILES) {
            console.log(
                `VISHNU_REACHED_PROFILE_LIMIT_${MAX_PROFILES}`
            );
            break;
        }

        await sleep(750);
    }

    if (savedProfileCount === 0) {
        throw new Error("ไม่สามารถเก็บข้อมูลสำหรับ Excel ได้");
    }

    await postToPython(
        "/complete",
        {
            status: "complete",
            total: savedProfileCount
        }
    );

    console.log(
        `VISHNU_ALL_PROFILE_FILES_SAVED_${savedProfileCount}`
    );

    return {
        status: "complete",
        total: savedProfileCount
    };
})()
"""

JAVASCRIPT = (
    JAVASCRIPT
    .replace("__MAX_PROFILES__", str(MAX_PROFILES))
    .replace("__TOTAL_QUESTIONS__", str(TOTAL_QUESTIONS))
    .replace("__SERVER_HOST__", LOCAL_SERVER_HOST)
    .replace("__SERVER_PORT__", str(LOCAL_SERVER_PORT))
)


# ============================================================
# Clipboard
# ============================================================

def set_clipboard(text: str) -> None:
    root = tk.Tk()
    root.withdraw()

    try:
        root.clipboard_clear()
        root.clipboard_append(text)
        root.update()
    finally:
        root.destroy()


def get_clipboard() -> str:
    root = tk.Tk()
    root.withdraw()

    try:
        root.update()
        return root.clipboard_get()
    except tk.TclError:
        return ""
    finally:
        root.destroy()


# ============================================================
# AHK และ Opera
# ============================================================

def create_ahk() -> AHK:
    ahk_path = Path(AHK_EXE_PATH)

    if ahk_path.is_file():
        return AHK(executable_path=str(ahk_path))

    return AHK()


def focus_opera(ahk: AHK) -> None:
    opera_window = ahk.win_get(
        title="ahk_exe opera.exe"
    )

    if opera_window is None:
        raise RuntimeError(
            "ไม่พบหน้าต่าง Opera ที่เปิดอยู่"
        )

    # ไม่ใช้ restore() เพราะอาจคืนหน้าต่างไปยังสถานะ Snap ครึ่งจอ
    opera_window.activate()

    time.sleep(0.5)

    # บังคับให้ Opera เต็มจอ
    opera_window.maximize()

    time.sleep(0.8)

    # Activate ซ้ำเพื่อให้แน่ใจว่าเป็นหน้าต่างด้านหน้า
    opera_window.activate()

    time.sleep(0.5)

    if not opera_window.is_active():
        raise RuntimeError(
            "ไม่สามารถโฟกัสหน้าต่าง Opera ได้"
        )


def toggle_console(ahk: AHK) -> None:
    # เปิดเฉพาะ Console ของ DevTools
    ahk.send("^+j")

    time.sleep(CONSOLE_OPEN_DELAY)


def run_javascript_in_console(ahk: AHK) -> None:
    # ตั้งค่า Clipboard เป็น JavaScript
    set_clipboard(JAVASCRIPT)

    time.sleep(0.3)

    # วาง JavaScript ลง Console
    ahk.send("^v")

    time.sleep(0.5)

    # รัน JavaScript
    ahk.send("{Enter}")

    time.sleep(JAVASCRIPT_RUN_DELAY)


# ============================================================
# Local server สำหรับรับข้อมูลจาก JavaScript
# ============================================================

class ProfileReceiver:
    def __init__(self) -> None:
        self.saved_count = 0
        self.complete_event = threading.Event()
        self.error: Exception | None = None
        self.server: ThreadingHTTPServer | None = None
        self.thread: threading.Thread | None = None

    def prepare_directory(self) -> None:
        TEMP_PROFILES_DIR.mkdir(parents=True, exist_ok=True)

        for json_file in TEMP_PROFILES_DIR.glob("profile_*.json"):
            json_file.unlink(missing_ok=True)

    def start(self) -> None:
        self.prepare_directory()
        receiver = self

        class RequestHandler(BaseHTTPRequestHandler):
            def _send_headers(
                self,
                status: int = 200,
                content_type: str = "application/json",
            ) -> None:
                self.send_response(status)
                self.send_header("Content-Type", content_type)
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header(
                    "Access-Control-Allow-Headers",
                    "Content-Type",
                )
                self.send_header(
                    "Access-Control-Allow-Methods",
                    "POST, OPTIONS",
                )
                self.end_headers()

            def do_OPTIONS(self) -> None:
                self._send_headers(204)

            def do_POST(self) -> None:
                try:
                    content_length = int(
                        self.headers.get("Content-Length", "0")
                    )
                    raw_body = self.rfile.read(content_length)
                    payload = json.loads(raw_body.decode("utf-8"))

                    if self.path == "/profile":
                        if not isinstance(payload, dict):
                            raise ValueError(
                                "ข้อมูล profile ต้องเป็น JSON Object"
                            )

                        receiver.saved_count += 1
                        profile_number = receiver.saved_count

                        payload["Saved At"] = datetime.now().strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )

                        profile_file = (
                            TEMP_PROFILES_DIR
                            / f"profile_{profile_number:04d}.json"
                        )

                        with profile_file.open(
                            mode="w",
                            encoding="utf-8",
                        ) as file:
                            json.dump(
                                payload,
                                file,
                                ensure_ascii=False,
                                indent=4,
                            )

                        response = {
                            "status": "saved",
                            "saved_count": profile_number,
                            "file": profile_file.name,
                        }

                    elif self.path == "/complete":
                        response = {
                            "status": "complete",
                            "saved_count": receiver.saved_count,
                        }
                        receiver.complete_event.set()

                    else:
                        self._send_headers(404)
                        self.wfile.write(
                            b'{"error":"not found"}'
                        )
                        return

                    self._send_headers(200)
                    self.wfile.write(
                        json.dumps(response).encode("utf-8")
                    )

                except Exception as error:
                    receiver.error = error
                    receiver.complete_event.set()

                    self._send_headers(500)
                    self.wfile.write(
                        json.dumps(
                            {"error": str(error)},
                            ensure_ascii=False,
                        ).encode("utf-8")
                    )

            def log_message(self, format: str, *args) -> None:
                return

        self.server = ThreadingHTTPServer(
            (LOCAL_SERVER_HOST, LOCAL_SERVER_PORT),
            RequestHandler,
        )

        self.thread = threading.Thread(
            target=self.server.serve_forever,
            daemon=True,
        )
        self.thread.start()

        print(
            "เปิด Local receiver แล้ว: "
            f"http://{LOCAL_SERVER_HOST}:{LOCAL_SERVER_PORT}"
        )

    def wait_for_completion(self) -> list[dict]:
        completed = self.complete_event.wait(
            timeout=RESULT_WAIT_TIMEOUT
        )

        if not completed:
            raise RuntimeError(
                "รอ JavaScript ส่งข้อมูลครบเกินกำหนด"
            )

        if self.error is not None:
            raise RuntimeError(
                f"Local receiver เกิดข้อผิดพลาด: {self.error}"
            )

        records = []

        for profile_file in sorted(
            TEMP_PROFILES_DIR.glob("profile_*.json")
        ):
            with profile_file.open(
                mode="r",
                encoding="utf-8",
            ) as file:
                record = json.load(file)

            if isinstance(record, dict):
                records.append(record)

        if not records:
            raise RuntimeError(
                "ไม่พบไฟล์ข้อมูลบุคคลใน temp_profiles"
            )

        return records

    def stop(self) -> None:
        if self.server is not None:
            self.server.shutdown()
            self.server.server_close()


# ============================================================
# เขียนและอ่านไฟล์ Temp
# ============================================================

def write_temp_file(data: list[dict]) -> None:
    try:
        with TEMP_FILE.open(
            mode="w",
            encoding="utf-8",
        ) as file:
            json.dump(
                data,
                file,
                ensure_ascii=False,
                indent=4,
            )

    except OSError as error:
        raise RuntimeError(
            f"ไม่สามารถเขียนไฟล์ Temp ได้:\n{TEMP_FILE}"
        ) from error

    print(f"เขียนข้อมูลลงไฟล์ Temp แล้ว: {TEMP_FILE}")


def read_temp_file() -> list[dict]:
    if not TEMP_FILE.exists():
        raise FileNotFoundError(
            f"ไม่พบไฟล์ Temp:\n{TEMP_FILE}"
        )

    try:
        with TEMP_FILE.open(
            mode="r",
            encoding="utf-8",
        ) as file:
            data = json.load(file)

    except json.JSONDecodeError as error:
        raise RuntimeError(
            "ข้อมูลในไฟล์ Temp ไม่ใช่ JSON ที่ถูกต้อง"
        ) from error

    except OSError as error:
        raise RuntimeError(
            f"ไม่สามารถอ่านไฟล์ Temp ได้:\n{TEMP_FILE}"
        ) from error

    if not isinstance(data, list):
        raise RuntimeError(
            "ข้อมูลในไฟล์ Temp ไม่ใช่ JSON Array"
        )

    if not all(isinstance(item, dict) for item in data):
        raise RuntimeError(
            "ข้อมูลบางรายการในไฟล์ Temp ไม่ใช่ JSON Object"
        )

    return data


# ============================================================
# Excel
# ============================================================

def download_image(image_url: str) -> bytes:
    request = urllib.request.Request(
        image_url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/149 Safari/537.36"
            )
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            return response.read()

    except Exception as error:
        raise RuntimeError(
            f"ดาวน์โหลด Profile Picture ไม่สำเร็จ:\n{image_url}"
        ) from error


def insert_profile_picture(
    sheet,
    row_number: int,
    column_number: int,
    image_url: str,
) -> None:
    if not image_url.startswith(("http://", "https://")):
        return

    image_bytes = download_image(image_url)

    # เปิดด้วย Pillow และแปลงเป็น PNG เพื่อให้ openpyxl รองรับแน่นอน
    with PILImage.open(io.BytesIO(image_bytes)) as source_image:
        source_image = source_image.convert("RGBA")

        # จำกัดขนาดรูป โดยรักษาอัตราส่วน
        max_width = 140
        max_height = 140

        source_image.thumbnail(
            (max_width, max_height),
            PILImage.Resampling.LANCZOS,
        )

        png_buffer = io.BytesIO()
        source_image.save(png_buffer, format="PNG")
        png_buffer.seek(0)

    excel_image = ExcelImage(png_buffer)

    excel_image.width = source_image.width
    excel_image.height = source_image.height

    cell_coordinate = sheet.cell(
        row=row_number,
        column=column_number,
    ).coordinate

    sheet.add_image(excel_image, cell_coordinate)

    # ความสูงของแถวใน Excel ใช้หน่วย point
    sheet.row_dimensions[row_number].height = max(
        110,
        source_image.height * 0.75,
    )

    # เก็บ buffer ไว้กับ object เพื่อป้องกันถูกปิดก่อน workbook.save()
    excel_image._vishnu_buffer = png_buffer


def append_to_excel(data: dict) -> int:
    # เก็บ URL รูปโปรไฟล์แยกไว้อีกคอลัมน์
    profile_picture_url = data.get("Profile Picture", "")

    if (
        isinstance(profile_picture_url, str)
        and profile_picture_url.startswith(("http://", "https://"))
    ):
        data["Profile Picture URL"] = profile_picture_url
    else:
        data["Profile Picture URL"] = ""

    preferred_headers = [
        "Question",
        "Profile Picture",
        "Profile Picture URL",
        "ชื่อเล่น",
        "จังหวัด",
        "โรงเรียนเดิม",
        "ช่องทางติดต่อ",
        "อาหารที่ชอบ",
        "สิ่งที่ชอบ",
        "ความสามารถ/งานอดิเรก",
        "สายการเรียน",
        "แท็กบุคลิก",
        "Page Title",
        "Source URL",
        "Saved At",
    ]

    if EXCEL_FILE.exists():
        workbook = load_workbook(EXCEL_FILE)

        if SHEET_NAME in workbook.sheetnames:
            sheet = workbook[SHEET_NAME]
        else:
            sheet = workbook.create_sheet(SHEET_NAME)

    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

    # อ่าน Header เดิม
    headers = []

    for cell in sheet[1]:
        if cell.value not in (None, ""):
            headers.append(str(cell.value))

    # จัดลำดับ Header โดยให้ Question เป็นคอลัมน์แรกเสมอ
    ordered_headers = []

    for header in preferred_headers:
        if header in data or header in headers:
            ordered_headers.append(header)

    for header in headers:
        if header not in ordered_headers:
            ordered_headers.append(header)

    for header in data:
        if header not in ordered_headers:
            ordered_headers.append(header)

    headers = ordered_headers

    # เขียน Header
    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # แถวใหม่
    next_row = sheet.max_row + 1

    if next_row < 2:
        next_row = 2

    # เขียนข้อมูล
    for column_number, header in enumerate(headers, start=1):
        value = data.get(header, "")

        cell = sheet.cell(
            row=next_row,
            column=column_number,
        )

        # Profile Picture จะใส่เป็นรูป ไม่ใส่ URL เป็นข้อความ
        if header == "Profile Picture":
            cell.value = ""

            if isinstance(value, str) and value:
                try:
                    insert_profile_picture(
                        sheet=sheet,
                        row_number=next_row,
                        column_number=column_number,
                        image_url=value,
                    )

                except Exception as error:
                    # ถ้าดาวน์โหลดหรือแทรกรูปไม่ได้ ให้เก็บ URL แทน
                    print(f"เตือน: แทรกรูปไม่สำเร็จ: {error}")

                    cell.value = value
                    cell.hyperlink = value
                    cell.style = "Hyperlink"

            continue

        cell.value = value

        if header == "Question":
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
        else:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        if (
            header in {"Profile Picture URL", "Source URL"}
            and isinstance(value, str)
            and value.startswith(("https://", "http://"))
        ):
            cell.hyperlink = value
            cell.style = "Hyperlink"

    # ตั้งความกว้างคอลัมน์
    for column_number, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_number)

        if header == "Question":
            sheet.column_dimensions[column_letter].width = 12

        elif header == "Profile Picture":
            # ประมาณ 140 pixels
            sheet.column_dimensions[column_letter].width = 22

        elif header in {
            "Profile Picture URL",
            "Source URL",
            "สิ่งที่ชอบ",
            "ความสามารถ/งานอดิเรก",
            "แท็กบุคลิก",
        }:
            sheet.column_dimensions[column_letter].width = 45

        else:
            sheet.column_dimensions[column_letter].width = 22

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    try:
        workbook.save(EXCEL_FILE)
    finally:
        workbook.close()

    return next_row


# ============================================================
# Main
# ============================================================

def main() -> None:
    receiver = ProfileReceiver()

    try:
        receiver.start()

        print("กำลังค้นหาหน้าต่าง Opera...")

        ahk = create_ahk()

        focus_opera(ahk)

        print("โฟกัสและขยาย Opera เต็มจอสำเร็จ")
        print("กำลังเปิด DevTools Console...")

        toggle_console(ahk)

        print("กำลังรัน JavaScript...")

        run_javascript_in_console(ahk)

        # ขั้นตอนที่ 1: รอ JavaScript ส่งไฟล์ข้อมูลรายคนมาครบ
        received_records = receiver.wait_for_completion()

        # ขั้นตอนที่ 2: รวมข้อมูลทั้งหมดลง TXT สำรอง
        write_temp_file(received_records)

        # ขั้นตอนที่ 3: อ่านข้อมูลทั้งหมดกลับจาก TXT
        records = read_temp_file()

        print("\nข้อมูลที่อ่านจากไฟล์ Temp")
        print("=" * 60)
        print(f"พบข้อมูลทั้งหมด {len(records)} คน")

        # ขั้นตอนที่ 4: เขียนทุกคนต่อท้าย Excel
        first_row = None
        last_row = None

        for index, record in enumerate(records, start=1):
            row_number = append_to_excel(record)

            if first_row is None:
                first_row = row_number

            last_row = row_number

            print(
                f"[{index}/{len(records)}] "
                f"เพิ่มข้อมูลลง Excel แถวที่ {row_number}"
            )

        print("=" * 60)
        print(
            f"เพิ่มข้อมูลทั้งหมด {len(records)} คน "
            f"ลง Excel แถวที่ {first_row}-{last_row}"
        )
        print(f"ไฟล์ Excel: {EXCEL_FILE}")
        print(f"ไฟล์ Temp: {TEMP_FILE}")

        if DELETE_TEMP_AFTER_SUCCESS:
            TEMP_FILE.unlink(missing_ok=True)
            print("ลบไฟล์ Temp แล้ว")

        toggle_console(ahk)

    except Exception as error:
        print(f"\nเกิดข้อผิดพลาด:\n{error}")
        raise

    finally:
        receiver.stop()


if __name__ == "__main__":
    main()